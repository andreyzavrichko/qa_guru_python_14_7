import datetime
import zipfile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
import pandas as pd


def test_archive_content():
    archive_path = 'resources/archive.zip'

    with zipfile.ZipFile(archive_path, 'r') as zf:
        assert 'employee.xlsx' in zf.namelist()
        assert 'report.pdf' in zf.namelist()
        assert 'sampleEcwid.csv' in zf.namelist()

        with zf.open('employee.xlsx') as file:
            wb = load_workbook(file)
            ws = wb.active
            assert ws['B2'].value == 'Васильков Петр Григорьевич'
            assert ws['C2'].value == 'Инженер 1-й категории'
            assert ws['D2'].value == datetime.datetime(2020, 2, 21, 0, 0)
            assert ws['E2'].value == 4

        with zf.open('report.pdf') as file:
            reader = PdfReader(file)
            page = reader.pages[0]
            text = page.extract_text()
            assert 'БАЛАНС 494669' in text, "Искомый текст не найден в PDF файле"

        with zf.open('sampleEcwid.csv') as file:
            df = pd.read_csv(file)
            assert df.shape == (1, 3)
            assert df.iloc[0]['name'] == 'Billabong'
            assert df.iloc[0]['sku'] == 'M115GTLE'
            assert df.iloc[0]['subtitle'] == 'Pre-order available'
