import datetime
import zipfile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
import pandas as pd

from script_os import ZIP_DIR


def test_xlsx():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zf:
        with zf.open('employee.xlsx') as xlsx_file:
            wb = load_workbook(xlsx_file)
            ws = wb.active
            assert ws['B2'].value == 'Васильков Петр Григорьевич'
            assert ws['C2'].value == 'Инженер 1-й категории'
            assert ws['D2'].value == datetime.datetime(2020, 2, 21, 0, 0)
            assert ws['E2'].value == 4


def test_pdf():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zf:
        with zf.open('report.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert 'БАЛАНС 494669' in text, "Искомый текст не найден в PDF файле"


def test_csv():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zf:
        with zf.open('sampleEcwid.csv') as file:
            df = pd.read_csv(file)
            assert df.shape == (1, 3)
            assert df.iloc[0]['name'] == 'Billabong'
            assert df.iloc[0]['sku'] == 'M115GTLE'
            assert df.iloc[0]['subtitle'] == 'Pre-order available'
